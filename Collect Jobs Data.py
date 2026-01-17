import pandas as pd
import json
import requests
from openpyxl import Workbook, load_workbook

api_url = "https://cf-courses-data.s3.us.cloud-object-storage.appdomain.cloud/IBM-DA0321EN-SkillsNetwork/labs/module%201/Accessing%20Data%20Using%20APIs/jobs.json"
# Number of Jobs for Python Technology 
def get_number_of_jobs_T(technology):
    response = requests.get(api_url)
    data= response.json()
    filtered_tech = []
    for tech in data:
        if tech["Key Skills"] == technology:
            filtered_tech.append(tech)
    number_of_jobs = len(filtered_tech)
    return technology, number_of_jobs

# print(get_number_of_jobs_T("Technical Support"))


'''# Number of Jobs for US Location
def get_number_of_jobs_L(location):
    # Fetch ALL jobs from the API (no params)
    response1 = requests.get(api_url)
    data = response1.json()
    # Filter the data manually by location  # filtered_jobs = [job for job in data if job["Location"] == location]
    filtered_jobs = []
    for job in data:
        if job["Location"] == location:
            filtered_jobs.append(job)
    
    number_of_jobs = len(filtered_jobs) 
    return location, number_of_jobs
print(get_number_of_jobs_L("Boston"))
'''

'''# Store the result in excel file_1 Location
locations = ["Los Angeles", "New York", "San Francisco", "Washington DC", "Seattle", "Austin", "Detroit"]
wb = Workbook()
ws = wb.active                              # Select Active Sheet
ws.title = "Job Postings"                   # Optional_ Rename active sheet
ws.append(["Location", "Number of Jobs"])   # Define Column Headers
#Loop through list
for city in locations:
    name, count = get_number_of_jobs_L(city)
    ws.append([name, count])

# Save the workbook as an excel file
wb.save("C:/Users/COM/OneDrive/Desktop/Hein Projects/python/Capstone Project/Lab/job-listings.xlsx")
print("File saved successfully!")
'''

# Job posting with language
languages = ['C#','C++','Java','JavaScript','Python','Scala','Oracle','SQL Server','MySQL Server','PostgreSQL','MongoDB']
wb = load_workbook("C:/Users/COM/OneDrive/Desktop/Hein Projects/python/Capstone Project/Lab/job-listings.xlsx")
ws=wb.create_sheet("CS_Language")
ws.append(["Language", "Number of Jobs"])
for language in languages:
    name, count = get_number_of_jobs_T(language)
    ws.append([name,count])
wb.save("C:/Users/COM/OneDrive/Desktop/Hein Projects/python/Capstone Project/Lab/job-listings.xlsx")
print("File saved successfully!")